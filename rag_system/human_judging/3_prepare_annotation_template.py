from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np
import pandas as pd


##################################################
# STEP 3: Stratified sample for human annotations
##################################################

def stratified_sample_with_category_validation(df, n_total=150, n_category_validation=60, seed=42):
    """
    Sample 150 examples:
    - All 150: answer quality metrics
    - Subset 60: category validation (20 per annotator)
    
    Category validation stratified by:
    - Dimension (QC: 20, AT: 20, LV: 20)
    - Granularity (balanced across coarse/medium/fine within RQ2)
    """
    
    np.random.seed(seed)
    samples = []
    
    # --- RQ1: Sample 50 (coarse-level, diverse dimensions) ---
    print("\n=== Sampling RQ1 ===")
    rq1_df = df[df['rq'] == 'RQ1'].copy()
    rq1_df = rq1_df[rq1_df['dimension'].isin(['QC', 'AT', 'LV'])]
    
    for dim in ['QC', 'AT', 'LV']:
        dim_df = rq1_df[rq1_df['dimension'] == dim]
        
        if len(dim_df) == 0:
            continue
        
        median_cs = dim_df['cosine_similarity'].median()
        high_perf = dim_df[dim_df['cosine_similarity'] >= median_cs]
        low_perf = dim_df[dim_df['cosine_similarity'] < median_cs]
        
        n_high = 8
        n_low = 8
        
        if len(high_perf) >= n_high:
            sample_high = high_perf.sample(n=n_high, random_state=seed+len(samples))
            samples.append(sample_high)
        else:
            samples.append(high_perf)
        
        if len(low_perf) >= n_low:
            sample_low = low_perf.sample(n=n_low, random_state=seed+len(samples))
            samples.append(sample_low)
        else:
            samples.append(low_perf)
    
    rq1_count = sum(len(s) for s in samples)
    print(f"RQ1 total sampled: {rq1_count}")
    
    # --- RQ2: Sample 78 (hierarchical) ---
    print("\n=== Sampling RQ2 ===")
    rq2_df = df[df['rq'] == 'RQ2'].copy()
    
    for dim in ['QC', 'AT', 'LV']:
        for gran in ['coarse', 'medium', 'fine']:
            gran_df = rq2_df[
                (rq2_df['dimension'] == dim) & 
                (rq2_df['granularity'] == gran)
            ]
            
            if len(gran_df) == 0:
                continue
            
            median_cs = gran_df['cosine_similarity'].median()
            high_perf = gran_df[gran_df['cosine_similarity'] >= median_cs]
            low_perf = gran_df[gran_df['cosine_similarity'] < median_cs]
            
            n_high = 5 if gran == 'fine' else 4
            n_low = 5 if gran == 'fine' else 4
            
            if len(high_perf) >= n_high:
                sample_high = high_perf.sample(n=n_high, random_state=seed+len(samples))
                samples.append(sample_high)
            else:
                samples.append(high_perf)
            
            if len(low_perf) >= n_low:
                sample_low = low_perf.sample(n=n_low, random_state=seed+len(samples))
                samples.append(sample_low)
            else:
                samples.append(low_perf)
    
    rq2_count = sum(len(s) for s in samples) - rq1_count
    print(f"RQ2 total sampled: {rq2_count}")
    
    # --- RQ3: Sample 25 ---
    print("\n=== Sampling RQ3 ===")
    rq3_df = df[df['rq'] == 'RQ3'].copy()
    
    if len(rq3_df) >= 24:
        rq3_sample = rq3_df.sample(n=24, random_state=seed)
        samples.append(rq3_sample)
    else:
        samples.append(rq3_df)
    
    # Combine
    final_sample = pd.concat(samples, ignore_index=True)
    
    if len(final_sample) > n_total:
        final_sample = final_sample.sample(n=n_total, random_state=seed)
    
    # --- SELECT SUBSET FOR CATEGORY VALIDATION (60 questions) ---
    # Stratify by dimension: 20 QC, 20 AT, 20 LV from RQ2 (fine-grained)
    category_validation_samples = []
    
    rq2_sample = final_sample[final_sample['rq'] == 'RQ2']
    
    for dim in ['QC', 'AT', 'LV']:
        # Prioritize fine-grained categories
        dim_fine = rq2_sample[(rq2_sample['dimension'] == dim) & (rq2_sample['granularity'] == 'fine')]
        
        # CHANGED: Ensure we get exactly 20 per dimension
        if len(dim_fine) >= 20:
            cat_val_sample = dim_fine.sample(n=20, random_state=seed+100)
        else:
            # Fill remaining from medium/coarse
            n_needed = 20 - len(dim_fine)
            dim_medium = rq2_sample[(rq2_sample['dimension'] == dim) & (rq2_sample['granularity'].isin(['medium', 'coarse']))]
            
            if len(dim_medium) >= n_needed:
                additional = dim_medium.sample(n=n_needed, random_state=seed+100)
            else:
                additional = dim_medium  # Take all available
            
            cat_val_sample = pd.concat([dim_fine, additional])
        
        category_validation_samples.append(cat_val_sample)
    
    category_validation_ids = pd.concat(category_validation_samples)['qa_id'].tolist()
    
    # Mark which questions need category validation
    final_sample['needs_category_validation'] = final_sample['qa_id'].isin(category_validation_ids)
    
    print(f"\n=== Final Sample: {len(final_sample)} questions ===")
    print(f"Category validation subset: {final_sample['needs_category_validation'].sum()}")
    
    return final_sample


######################
df_all = pd.read_json('./annotate_files/combined_rq1_rq2_rq3_with_metrics.jsonl', lines=True)  # Load combined data with metrics

# Generate sample
annotation_df = stratified_sample_with_category_validation(df_all, n_total=150, n_category_validation=60, seed=42)

print("\n=== Distribution ===")
print("By RQ:")
print(annotation_df['rq'].value_counts())
print("\nCategory validation by dimension:")
cat_val = annotation_df[annotation_df['needs_category_validation']]
print(cat_val['dimension'].value_counts())


# Verify distribution
print("\n=== Sample Distribution ===")
print("\nBy RQ:")
print(annotation_df['rq'].value_counts().sort_index())

print("\nBy Dimension:")
print(annotation_df['dimension'].value_counts())

print("\nBy Granularity (RQ2 only):")
rq2_sample = annotation_df[annotation_df['rq'] == 'RQ2']
print(rq2_sample['granularity'].value_counts())

print("\nBy Dimension × Granularity (RQ2):")
print(rq2_sample.groupby(['dimension', 'granularity']).size())

print("\nPerformance distribution (CS quartiles):")
print(annotation_df['cosine_similarity'].describe())


##################################
# STEP 4: Prepare templates with category validation -- for human judges
##################################

# Fill missing granularity
annotation_df['granularity'] = annotation_df['granularity'].fillna('N/A')

# Prepare annotation template
annotation_template = annotation_df[[
    'qa_id',
    'question',
    'reference_answer',
    'generated_answer',
    'context',
    'rq',
    'dimension',
    'granularity',
    'category_name',
    'user_expertise',
    'cosine_similarity',
    'map_score',
    'needs_category_validation'
]].copy()

# Extract source document
annotation_template['source_document'] = annotation_template['context'].apply(
    lambda x: x[0] if isinstance(x, list) and len(x) > 0 else ''
)
annotation_template = annotation_template.drop('context', axis=1)

# Truncate long documents
annotation_template['source_document'] = annotation_template['source_document'].apply(
    # lambda x: x
    # OLD: truncated content. Adjusted to keep full content for annotation.
    lambda x: x[:1500] + '...[truncated]' if len(str(x)) > 1500 else x
)

# Add annotation columns
annotation_template['answer_correctness'] = ''
annotation_template['hallucination_score'] = ''
annotation_template['question_answerability'] = ''
annotation_template['category_validation'] = ''  # Only for subset
annotation_template['category_confidence'] = ''  # Only for subset
annotation_template['notes'] = ''

# Round metrics
annotation_template['cosine_similarity'] = annotation_template['cosine_similarity'].round(3)
annotation_template['map_score'] = annotation_template['map_score'].round(3)

print(f" Template prepared: {len(annotation_template)} rows")






##########################################################
# Step 5: Create Shared Questions + Split with Overlap
##########################################################

# Select 30 shared questions (for inter-annotator agreement)
def select_shared_questions(df, n_shared=30):
    shared = []
    
    # RQ1: 10 questions
    rq1 = df[df['rq'] == 'RQ1'].copy()
    if len(rq1) >= 10:
        shared.append(rq1.sample(n=10, random_state=42))
    
    # RQ2: 15 questions
    rq2 = df[df['rq'] == 'RQ2'].copy()
    if len(rq2) >= 15:
        shared.append(rq2.sample(n=15, random_state=42))
    
    # RQ3: 5 questions
    rq3 = df[df['rq'] == 'RQ3'].copy()
    if len(rq3) >= 5:
        shared.append(rq3.sample(n=5, random_state=42))
    
    shared_df = pd.concat(shared, ignore_index=True)
    return shared_df

shared_questions = select_shared_questions(annotation_template, n_shared=30)
print(f" Selected {len(shared_questions)} shared questions")

# Get remaining questions (after removing shared)
remaining_questions = annotation_template[~annotation_template['qa_id'].isin(shared_questions['qa_id'])]

# Separate shared into: cat_val vs non-cat_val
shared_cat_val = shared_questions[shared_questions['needs_category_validation'] == True]
shared_no_cat_val = shared_questions[shared_questions['needs_category_validation'] == False]

# Separate remaining into: cat_val vs non-cat_val
remaining_cat_val = remaining_questions[remaining_questions['needs_category_validation'] == True]
remaining_no_cat_val = remaining_questions[remaining_questions['needs_category_validation'] == False]

# Each annotator gets:
# - All shared (30 questions)
# - 40 unique questions with balanced cat_val distribution

# Split remaining cat_val questions evenly (should be ~24 remaining cat_val / 3 = 8 each)
n_cat_val_per_annotator = len(remaining_cat_val) // 3

unique_cat_val_1 = remaining_cat_val.iloc[0:n_cat_val_per_annotator].copy()
unique_cat_val_2 = remaining_cat_val.iloc[n_cat_val_per_annotator:2*n_cat_val_per_annotator].copy()
unique_cat_val_3 = remaining_cat_val.iloc[2*n_cat_val_per_annotator:].copy()

# Split remaining non-cat_val to fill to 40 unique each
n_regular_per_annotator = 40 - n_cat_val_per_annotator

unique_regular_1 = remaining_no_cat_val.iloc[0:n_regular_per_annotator].copy()
unique_regular_2 = remaining_no_cat_val.iloc[n_regular_per_annotator:2*n_regular_per_annotator].copy()
unique_regular_3 = remaining_no_cat_val.iloc[2*n_regular_per_annotator:3*n_regular_per_annotator].copy()

# Combine for each annotator
annotator_1 = pd.concat([shared_questions, unique_cat_val_1, unique_regular_1], ignore_index=True)
annotator_2 = pd.concat([shared_questions, unique_cat_val_2, unique_regular_2], ignore_index=True)
annotator_3 = pd.concat([shared_questions, unique_cat_val_3, unique_regular_3], ignore_index=True)

# Sort by needs_category_validation (TRUE first) - yellow rows at top
annotator_1 = annotator_1.sort_values('needs_category_validation', ascending=False).reset_index(drop=True)
annotator_2 = annotator_2.sort_values('needs_category_validation', ascending=False).reset_index(drop=True)
annotator_3 = annotator_3.sort_values('needs_category_validation', ascending=False).reset_index(drop=True)

print(f"\n Annotator batches created:")
for i, batch in enumerate([annotator_1, annotator_2, annotator_3], 1):
    n_cat_val = batch['needs_category_validation'].sum()
    print(f"  Annotator {i}: {len(batch)} total, {n_cat_val} category validation")

print(f"\n Annotator batches created:")
print(f"  Annotator 1: {len(annotator_1)} questions")
print(f"  Annotator 2: {len(annotator_2)} questions")
print(f"  Annotator 3: {len(annotator_3)} questions")


##########################################################
# Step 6: Format Excel Files with Conditional Formatting
##########################################################

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def format_annotation_excel(filename, annotator_num):
    """
    Apply Excel formatting:
    - Hide metadata columns (CS, MAP)
    - Highlight category validation rows
    - Add data validation dropdowns
    - Freeze panes
    """
    
    wb = load_workbook(filename)
    ws = wb.active
    
    # Column indices (adjust based on our actual column order)
    col_mapping = {
        'qa_id': 1,
        'rq': 2,
        'dimension': 3,
        'granularity': 4,
        'category_name': 5,
        'user_expertise': 6,
        'source_document': 7,
        'question': 8,
        'reference_answer': 9,
        'generated_answer': 10,
        'cosine_similarity': 11,
        'map_score': 12,
        'needs_category_validation': 13,
        'answer_correctness': 14,
        'hallucination_score': 15,
        'question_answerability': 16,
        'category_validation': 17,
        'category_confidence': 18,
        'notes': 19
    }
    
    # 1. Hide metadata columns (CS, MAP, needs_category_validation)
    for col_name in ['cosine_similarity', 'map_score', 'needs_category_validation']:
        col_idx = col_mapping[col_name]
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True
    
    # 2. Format header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 3. Highlight rows needing category validation
    cat_val_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    for row_idx in range(2, ws.max_row + 1):
        needs_val = ws.cell(row=row_idx, column=col_mapping['needs_category_validation']).value
        
        if needs_val == True or needs_val == 'TRUE' or needs_val == 1:
            # Highlight entire row
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = cat_val_fill
    
    # 4. Set column widths
    ws.column_dimensions[get_column_letter(col_mapping['qa_id'])].width = 12
    ws.column_dimensions[get_column_letter(col_mapping['dimension'])].width = 12
    ws.column_dimensions[get_column_letter(col_mapping['category_name'])].width = 25
    ws.column_dimensions[get_column_letter(col_mapping['source_document'])].width = 60
    ws.column_dimensions[get_column_letter(col_mapping['question'])].width = 50
    ws.column_dimensions[get_column_letter(col_mapping['reference_answer'])].width = 60
    ws.column_dimensions[get_column_letter(col_mapping['generated_answer'])].width = 60
    ws.column_dimensions[get_column_letter(col_mapping['answer_correctness'])].width = 15
    ws.column_dimensions[get_column_letter(col_mapping['hallucination_score'])].width = 15
    ws.column_dimensions[get_column_letter(col_mapping['question_answerability'])].width = 18
    ws.column_dimensions[get_column_letter(col_mapping['category_validation'])].width = 30
    ws.column_dimensions[get_column_letter(col_mapping['notes'])].width = 40
    
    # 5. Wrap text for long columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_name in ['source_document', 'question', 'reference_answer', 'generated_answer', 'notes']:
            cell = row[col_mapping[col_name] - 1]
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # 6. Freeze panes (freeze first row + first 6 columns for scrolling)
    ws.freeze_panes = 'G2'
    
    # 7. Add data validation (dropdowns) for rating columns
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # 7. Add DYNAMIC dropdowns for category validation (dimension-specific)
    from openpyxl.worksheet.datavalidation import DataValidation

    # Category options by dimension
    qc_options = '"extractive_span,entity_extraction,multi_span_aggregation,paraphrasing_required,single_hop_inference,multi_hop_reasoning,comparative_synthesis,comparative_synthesis_concepts"'

    at_options = '"entity_extraction,phrase_extraction,unordered_list,ordered_sequence,condensed_summary,sentence_extraction,analytical_synthesis,explanatory_synthesis"'

    lv_options = '"verbatim_terminology,high_lexical_overlap,moderate_lexical_overlap,moderate_low_lexical_overlap,low_lexical_overlap,synonym_based_rephrase,domain_shift_terminology,abstraction_level_shift"'

    # Apply dimension-specific dropdown to each row
    for row_idx in range(2, ws.max_row + 1):
        dimension = ws.cell(row=row_idx, column=col_mapping['dimension']).value
        needs_val = ws.cell(row=row_idx, column=col_mapping['needs_category_validation']).value
        
        if needs_val == True or needs_val == 'TRUE' or needs_val == 1:
            # Determine which dropdown to use
            if dimension == 'QC':
                formula = qc_options
            elif dimension == 'AT':
                formula = at_options
            elif dimension == 'LV':
                formula = lv_options
            else:
                continue
            
            # Create dimension-specific dropdown for this row only
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(dv)
            cell_ref = f'{get_column_letter(col_mapping["category_validation"])}{row_idx}'
            dv.add(cell_ref)
            
    # Answer correctness: 0-3
    dv_correctness = DataValidation(type="list", formula1='"0,1,2,3"', allow_blank=True)
    dv_correctness.error = 'Please enter 0, 1, 2, or 3'
    ws.add_data_validation(dv_correctness)
    dv_correctness.add(f'{get_column_letter(col_mapping["answer_correctness"])}2:{get_column_letter(col_mapping["answer_correctness"])}{ws.max_row}')
    
    # Hallucination: 0-2
    dv_hallucination = DataValidation(type="list", formula1='"0,1,2"', allow_blank=True)
    dv_hallucination.error = 'Please enter 0, 1, or 2'
    ws.add_data_validation(dv_hallucination)
    dv_hallucination.add(f'{get_column_letter(col_mapping["hallucination_score"])}2:{get_column_letter(col_mapping["hallucination_score"])}{ws.max_row}')
    
    # Question answerability: 0-3
    dv_answerability = DataValidation(type="list", formula1='"0,1,2,3"', allow_blank=True)
    dv_answerability.error = 'Please enter 0, 1, 2, or 3'
    ws.add_data_validation(dv_answerability)
    dv_answerability.add(f'{get_column_letter(col_mapping["question_answerability"])}2:{get_column_letter(col_mapping["question_answerability"])}{ws.max_row}')
    
    # Category confidence: High/Medium/Low
    dv_confidence = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(dv_confidence)
    dv_confidence.add(f'{get_column_letter(col_mapping["category_confidence"])}2:{get_column_letter(col_mapping["category_confidence"])}{ws.max_row}')
    
    wb.save(filename)
    print(f" Formatted: {filename}")

# Create Excel files for annotators
for i, batch in enumerate([annotator_1, annotator_2, annotator_3], 1):
    # Columns for annotator view (remove CS/MAP from view, keep in file hidden)
    annotator_columns = [
        'qa_id', 'rq', 'dimension', 'granularity', 'category_name', 'user_expertise',
        'source_document', 'question', 'reference_answer', 'generated_answer',
        'cosine_similarity', 'map_score', 'needs_category_validation',  # Will be hidden
        'answer_correctness', 'hallucination_score', 'question_answerability',
        'category_validation', 'category_confidence', 'notes'
    ]
    
    batch_export = batch[annotator_columns].copy()

    # Sort: category validation rows first
    batch_export = batch_export.sort_values('needs_category_validation', ascending=False).reset_index(drop=True)
        
    # TODO REAL: filename = f'./annotate_files/annotation_batch_annotator_{i}_HYBRID.xlsx'
    filename = f'./annotate_files/trunc_annotation_batch_annotator_{i}_HYBRID.xlsx'# For truncated.
    
    batch_export.to_excel(filename, index=False, engine='openpyxl')
    
    # Apply formatting
    format_annotation_excel(filename, i)
    
    print(f"\nAnnotator {i}:")
    print(f"  Total questions: {len(batch)}")
    print(f"  Category validation: {batch['needs_category_validation'].sum()}")

# Save shared questions list
shared_questions[['qa_id', 'question', 'dimension']].to_csv('./annotate_files/trunc_shared_questions_list.csv', index=False)

print("\n" + "="*60)
print(" All files created!")
print("\nFiles:")
print("  - annotation_batch_annotator_1_HYBRID.xlsx")
print("  - annotation_batch_annotator_2_HYBRID.xlsx")
print("  - annotation_batch_annotator_3_HYBRID.xlsx")
print("  - shared_questions_list.csv")



####################################################
# Step 7: Create Category Options Reference Sheet
####################################################

# Create category options for each dimension
category_options = {
    'QC': [
        'extractive_span',
        'entity_extraction',
        'multi_span_aggregation',
        'paraphrasing_required',
        'single_hop_inference',
        'multi_hop_reasoning',
        'comparative_synthesis',
        'comparative_synthesis_concepts'
    ],
    'AT': [
        'entity_extraction',
        'phrase_extraction',
        'unordered_list',
        'ordered_sequence',
        'condensed_summary',
        'sentence_extraction',
        'analytical_synthesis',
        'explanatory_synthesis'
    ],
    'LV': [
        'verbatim_terminology',
        'high_lexical_overlap',
        'moderate_lexical_overlap',
        'moderate_low_lexical_overlap',
        'low_lexical_overlap',
        'synonym_based_rephrase',
        'domain_shift_terminology',
        'abstraction_level_shift'
    ]
}

# Create category definitions reference
category_reference = """
# CATEGORY VALIDATION REFERENCE SHEET

For questions highlighted in YELLOW, you must also validate the category assignment.

## How to Validate Categories

1. Read the question's assigned category (in column 'category_name')
2. Based on the dimension (QC/AT/LV), review the options below
3. In the 'category_validation' column, enter which category BEST fits the question
4. In 'category_confidence' column, select: High / Medium / Low

---

## Question Complexity (QC) Categories

**extractive_span** - Copying exact text/phrase from document
- Example: "What is the molecular weight?" → Answer: "180.158 g/mol" (verbatim)

**entity_extraction** - Extracting a specific fact (name, date, number)
- Example: "When was the company founded?" → Answer: "1995"

**multi_span_aggregation** - Combining multiple separate facts from document
- Example: "What are the three main symptoms?" → Answer requires listing 3+ items

**paraphrasing_required** - Answer needs rewording of document text
- Example: Document says "extremely dangerous"; answer should say "very hazardous"

**single_hop_inference** - Simple reasoning from one section
- Example: "Why did sales increase?" → Document has revenue data + market info

**multi_hop_reasoning** - Combining information from multiple sections
- Example: "How did X affect Y?" → Requires connecting cause from section A to effect in section B

**comparative_synthesis** - Comparing two entities
- Example: "How does Product A differ from Product B?"

**comparative_synthesis_concepts** - Comparing abstract ideas/concepts
- Example: "What's the difference between democracy and autocracy?"

---

## Answer Type (AT) Categories

**entity_extraction** - Short factual answer (1-3 words)
- Example answer: "Napoleon Bonaparte" or "1812"

**phrase_extraction** - Short phrase (4-10 words)
- Example answer: "the Battle of Waterloo in Belgium"

**unordered_list** - Multiple items, order doesn't matter
- Example answer: "apples, oranges, bananas"

**ordered_sequence** - Multiple items in specific order
- Example answer: "First, preheat oven. Second, mix ingredients. Third, bake."

**condensed_summary** - Brief synthesis (1-2 sentences)
- Example answer: "The study found that X causes Y in 80% of cases."

**sentence_extraction** - Pulling one full sentence from document
- Example answer: Exact sentence from document

**analytical_synthesis** - Analysis combining multiple points
- Example answer: "The data suggests X because of factors A, B, and C."

**explanatory_synthesis** - Detailed explanation requiring synthesis
- Example answer: Multi-sentence explanation of how/why something works

---

## Linguistic Variation (LV) Categories

**verbatim_terminology** - Uses EXACT words from document
- Question uses same phrase as document word-for-word

**high_lexical_overlap** - Uses mostly document words (80%+ overlap)
- Question uses most key terms from document

**moderate_lexical_overlap** - Uses some document words (50-80% overlap)
- Question uses some key terms, some paraphrased

**moderate_low_lexical_overlap** - Uses few document words (20-50% overlap)
- Question has partial overlap with document terminology

**low_lexical_overlap** - Uses very few document words (<20% overlap)
- Question mostly uses different words than document

**synonym_based_rephrase** - Uses synonyms of document terms
- Document: "dangerous"; Question: "hazardous"
- Document: "increase"; Question: "rise"

**domain_shift_terminology** - Uses terms from different domain
- Document (medical): "myocardial infarction"; Question (layman): "heart attack"

**abstraction_level_shift** - Different conceptual level
- Document (specific): "iPhone 14"; Question (general): "smartphones"

---

## Instructions

For YELLOW-HIGHLIGHTED rows only:
1. Read the question and document
2. Look at the assigned category in 'category_name' column
3. Review the options above for that dimension
4. Enter your selection in 'category_validation' column (exact category name)
5. Rate confidence: High / Medium / Low

For non-highlighted rows: Leave 'category_validation' and 'category_confidence' BLANK.
"""

with open('category_validation_reference.txt', 'w') as f:
    f.write(category_reference)

print("Category reference sheet created: category_validation_reference.txt")

##########################################
# Step 8: Create Annotator Instructions
##########################################


instructions = """
# RAG Benchmark Annotation Instructions - HYBRID VERSION

## Overview
You will evaluate **70 question-answer pairs** in two ways:

1. **ALL 70 questions:** Rate answer quality (correctness, hallucination, answerability)
2. **~20 questions (HIGHLIGHTED IN YELLOW):** Also validate category assignment

**Total estimated time:** 25-30 minutes

---

## Part 1: Answer Quality (ALL 70 Questions)

### Rating 1: Answer Correctness (0-3)
Rate the `reference_answer` against the `source_document`:

- **3** = Correct & complete - fully addresses question based on document
- **2** = Mostly correct - minor details missing or small inaccuracies
- **1** = Partially correct - significant errors or omissions
- **0** = Incorrect/hallucinated - contradicts document or invents facts

### Rating 2: Hallucination Score (0-2)
Does the answer contain information NOT in the source document?

- **0** = No hallucination - all info grounded in document
- **1** = Minor hallucination - adds plausible inference/world knowledge (not contradictory)
- **2** = Major hallucination - invents specific facts not present in document

### Rating 3: Question Answerability (0-3)
Can the question be answered using the source document?

- **3** = Fully answerable - document contains all info needed
- **2** = Mostly answerable - document has most info but missing some details
- **1** = Partially answerable - document only hints at answer
- **0** = Unanswerable - document doesn't contain relevant info

---

## Part 2: Category Validation (YELLOW ROWS ONLY - ~20 questions)

For rows **highlighted in yellow**, you must also validate the category assignment.

### Rating 4: Category Validation
1. Check the `dimension` column (QC / AT / LV)
2. Look at the assigned `category_name`
3. Open the file **category_validation_reference.txt** to see all category options
4. In the `category_validation` column, type which category BEST fits the question
   - This might be the same as the assigned category (good!)
   - Or it might be different (that's OK, we want your honest assessment)

### Rating 5: Category Confidence
Select: **High** / **Medium** / **Low**

**IMPORTANT:** For non-highlighted rows, leave category validation columns BLANK.

---

## Examples

### Example 1: Answer Quality Only (White Row)

**Source Document:** "Aspirin has a molecular weight of 180.158 g/mol..."

**Question:** "What is the molecular weight of aspirin?"
**Reference Answer:** "180.158 g/mol"

**Your Ratings:**
- Answer Correctness: **3** (correct & complete)
- Hallucination: **0** (all from document)
- Question Answerability: **3** (fully answerable)
- Category Validation: **[LEAVE BLANK - not highlighted]**

---

### Example 2: With Category Validation (Yellow Row)

**Source Document:** "Palmer amaranth is ranked as the most difficult weed to control in the U.S. It can produce up to a million seeds per plant per season and has stems so tough they damage farm equipment."

**Question:** "What makes Palmer amaranth a nightmare for farmers?"
**Assigned Category:** synonym_substitution (LV)
**Reference Answer:** "It's difficult to control, produces a million seeds, and damages equipment."

**Your Ratings:**
- Answer Correctness: **2** (mostly correct, slightly incomplete)
- Hallucination: **0** (all from document)
- Question Answerability: **3** (fully answerable)
- Category Validation: **synonym_substitution** (document says "difficult weed to control", question says "nightmare" - that's synonym substitution! )
- Category Confidence: **High**

---

### Example 3: Category Mismatch (Yellow Row)

**Question:** "How many seeds can one plant produce?"
**Assigned Category:** multi_hop_reasoning (QC)

**Your Assessment:** This is simple extraction (finding one number), NOT multi-hop reasoning!

**Your Ratings:**
- Category Validation: **entity_extraction**
- Category Confidence: **High**

---

## Tips

 **Answer quality first** - Do all 70 questions for answer ratings, then go back to yellow rows for category validation
 **Use the reference sheet** - Keep category_validation_reference.txt open while working
 **Be honest** - If the assigned category seems wrong, say so! That's valuable data
 **When in doubt** - Select "Medium" or "Low" confidence
---

Thank you!
"""

with open('./annotate_files/annotator_instructions_HYBRID.txt', 'w') as f:
    f.write(instructions)

print(" Instructions created: annotator_instructions_HYBRID.txt")




